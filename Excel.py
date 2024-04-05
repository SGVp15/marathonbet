import openpyxl

from config import excel_file
from match import Match


def add_to_excel_file(matchs: [Match]):
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active

    row = worksheet.max_row
    row = 1
    # Create a list of values to write to the Excel file
    for m in matchs:
        # m.country
        # m.championship
        # m.command_first_name
        # m.command_second_name
        # m.value_1_col
        # m.value_2_col
        # m.match_total_first_team_1p5_value
        # m.match_total_second_team_1p5_value
        # m.goals_dict
        for values_to_write in m.get_match_for_excel():
            row += 1
            # Write the numbers to the worksheet
            for i, v in enumerate(values_to_write):
                worksheet.cell(row=row, column=i + 1, value=str(v))

    # Save the workbook to a file
    workbook.save(excel_file)
